# ui/views/tools.py
import streamlit as st
import pandas as pd
import numpy as np
from scipy.stats import norm
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def render_tools():
    st.markdown("### 🛠️ Boîte à Outils Supply Chain")
    st.markdown("Passez de la théorie à l'application métier. Calculez, simulez et exportez.")

    tabs = st.tabs([
        "🛡️ Stock de Sécurité", 
        "💰 Coût Complet (Landed Cost)", 
        "📦 Poids Volumétrique",
        "🚢 Sélecteur Incoterms",
        "📊 ABC-XYZ & Formules"
    ])

    with tabs[0]:
        render_safety_stock_calculator()

    with tabs[1]:
        render_landed_cost_calculator()

    with tabs[2]:
        render_volumetric_calculator()
    
    with tabs[3]:
        render_incoterm_selector()

    with tabs[4]:
        render_templates_section()

# --- 1. STOCK DE SÉCURITÉ ---
def render_safety_stock_calculator():
    st.subheader("🛡️ Calculateur de Stock de Sécurité (Modèle de King)")
    st.info("Ce modèle combine l'incertitude de la demande ET la fiabilité du délai fournisseur.")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**📁 Demande**")
        avg_demand = st.number_input("Conso moyenne / jour", value=100.0, step=1.0, key="ss_avg_d")
        std_demand = st.number_input("Écart-type conso", value=20.0, step=1.0, key="ss_std_d")
        
        st.markdown("**🚚 Fournisseur**")
        lead_time = st.number_input("Délai moyen (jours)", value=10.0, step=1.0, key="ss_lt")
        std_lead_time = st.number_input("Écart-type délai (jours)", value=2.0, step=0.5, key="ss_std_lt")

    with col2:
        st.markdown("**🎯 Service**")
        service_level = st.slider("Taux de Service cible (%)", 80.0, 99.9, 95.0, 0.1, key="ss_sl")
        z_score = norm.ppf(service_level / 100)
        st.metric("Coefficient Z", f"{z_score:.2f}")

    combined_std = np.sqrt(lead_time * (std_demand ** 2) + (avg_demand ** 2) * (std_lead_time ** 2))
    safety_stock = z_score * combined_std

    st.markdown("---")
    c1, c2, c3 = st.columns(3)
    c1.metric("Stock de Sécurité", f"{int(np.ceil(safety_stock))} u.")
    c2.metric("Point de Commande", f"{int(np.ceil(avg_demand * lead_time + safety_stock))} u.")
    c3.metric("Stock Moyen (SS)", f"{int(np.ceil(safety_stock))} u.")

    if st.button("📥 Générer Excel Stock de Sécurité", use_container_width=True):
        output = create_ss_excel(avg_demand, std_demand, lead_time, std_lead_time, service_level, safety_stock, z_score)
        st.download_button("💾 Télécharger SS.xlsx", output, "MentorSC_Stock_Securite.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- 2. LANDED COST (COÛT COMPLET) ---
def render_landed_cost_calculator():
    st.subheader("💰 Comparateur de Coût Complet (Landed Cost)")
    st.write("Comparez deux sources d'approvisionnement en incluant les coûts cachés.")

    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 🇨🇳 Source A (Ex: Grand Import)")
        price_a = st.number_input("Prix unitaire (Ex-Works)", value=10.0, key="lc_p_a")
        freight_a = st.number_input("Fret unitaire", value=2.5, key="lc_f_a")
        customs_a = st.number_input("Droits de douane (%)", value=6.5, key="lc_c_a")
        lead_time_a = st.number_input("Temps de transport (jours)", value=45, key="lc_lt_a")

    with col2:
        st.markdown("### 🇪🇺 Source B (Ex: Proche Import)")
        price_b = st.number_input("Prix unitaire (Ex-Works)", value=14.0, key="lc_p_b")
        freight_b = st.number_input("Fret unitaire", value=0.8, key="lc_f_b")
        customs_b = st.number_input("Droits de douane (%)", value=0.0, key="lc_c_b")
        lead_time_b = st.number_input("Temps de transport (jours)", value=5, key="lc_lt_b")

    st.markdown("---")
    cost_capital = st.slider("Taux de possession annuel (%)", 5.0, 30.0, 15.0, 1.0) / 100

    # Calculs
    def calc_landed(p, f, c, lt):
        duties = (p + f) * (c / 100)
        # Coût financier du stock en transit
        financial_cost = (p + f + duties) * cost_capital * (lt / 365)
        total = p + f + duties + financial_cost
        return total, duties, financial_cost

    total_a, duty_a, fin_a = calc_landed(price_a, freight_a, customs_a, lead_time_a)
    total_b, duty_b, fin_b = calc_landed(price_b, freight_b, customs_b, lead_time_b)

    res1, res2 = st.columns(2)
    res1.metric("Total Landed A", f"{total_a:.2f} €", f"{((total_a/total_b)-1)*100:+.1f}% vs B")
    res2.metric("Total Landed B", f"{total_b:.2f} €", f"{((total_b/total_a)-1)*100:+.1f}% vs A")

    with st.expander("📊 Détail des coûts unitaires"):
        st.write(pd.DataFrame({
            "Poste de coût": ["Prix Achat", "Fret", "Douane", "Portage Financier", "TOTAL"],
            "Source A": [f"{price_a}€", f"{freight_a}€", f"{duty_a:.2f}€", f"{fin_a:.2f}€", f"{total_a:.2f}€"],
            "Source B": [f"{price_b}€", f"{freight_b}€", f"{duty_b:.2f}€", f"{fin_b:.2f}€", f"{total_b:.2f}€"]
        }))

# --- 3. POIDS VOLUMÉTRIQUE ---
def render_volumetric_calculator():
    st.subheader("📦 Calculateur de Poids Volumétrique")
    st.write("Le transporteur facture au plus élevé entre le poids réel et le volume.")

    c1, c2, c3 = st.columns(3)
    length = c1.number_input("Longueur (cm)", 1.0, 300.0, 60.0)
    width = c2.number_input("Largeur (cm)", 1.0, 300.0, 40.0)
    height = c3.number_input("Hauteur (cm)", 1.0, 300.0, 40.0)
    
    real_weight = st.number_input("Poids Réel (kg)", 0.1, 1000.0, 15.0)
    mode = st.radio("Mode de transport", ["Aérien (1:6)", "Route (1:3)", "Maritime (1:1)"], horizontal=True)

    ratios = {"Aérien (1:6)": 6000, "Route (1:3)": 3000, "Maritime (1:1)": 1000}
    ratio = ratios[mode]
    
    vol_weight = (length * width * height) / (ratio if mode != "Maritime (1:1)" else 1000)
    if mode == "Maritime (1:1)": vol_weight = (length * width * height) / 1000000 * 1000 # 1m3 = 1000kg

    st.markdown("---")
    res_c1, res_c2 = st.columns(2)
    res_c1.metric("Poids Volumétrique", f"{vol_weight:.2f} kg")
    
    is_vol = vol_weight > real_weight
    taxable = max(vol_weight, real_weight)
    
    color = "red" if is_vol else "green"
    st.markdown(f"""
    <div style="padding:20px; border-radius:10px; border:2px solid {color}; text-align:center;">
        <h3 style="color:{color};">Poids Taxable : {taxable:.2f} kg</h3>
        <p>{"⚠️ Attention : Vous payez pour du volume (votre colis est 'léger')." if is_vol else "✅ Optimal : Vous payez pour le poids réel."}</p>
    </div>
    """, unsafe_allow_html=True)

# --- 4. INCOTERMS ---
def render_incoterm_selector():
    st.subheader("🚢 Sélecteur d'Incoterm Idéal")
    st.write("Répondez à 3 questions pour trouver l'Incoterm adapté à votre stratégie.")

    q1 = st.toggle("Je veux maîtriser le coût et le choix du transporteur international")
    q2 = st.toggle("Je veux que le fournisseur gère toutes les formalités de douane import")
    q3 = st.toggle("Je veux limiter mes risques au maximum (livraison chez moi)")

    st.markdown("---")
    if q3 or q2:
        st.success("🎯 Recommandation : **DDP (Delivered Duty Paid)**")
        st.write("Le fournisseur s'occupe de tout jusqu'à votre porte. Confort maximal, mais coût caché probable.")
    elif q1:
        st.success("🎯 Recommandation : **FOB (Free On Board)** ou **FCA**")
        st.write("Vous maîtrisez la chaîne logistique dès le départ du pays d'origine. Idéal pour optimiser les coûts.")
    else:
        st.info("🎯 Recommandation : **EXW (Ex-Works)**")
        st.write("Le fournisseur met juste à disposition. Attention : vous gérez la douane export dans un pays étranger !")

# --- 5. ABC-XYZ & TEMPLATES ---
def render_templates_section():
    st.subheader("📊 Segmentation & Formules")
    
    with st.expander("📐 Aide à la décision ABC-XYZ"):
        st.write("""
        | Catégorie | XYZ (Prévisibilité) | Stratégie Recommandée |
        | :--- | :--- | :--- |
        | **AX** | Stable | Automatisation, Flux tendu, Stock bas. |
        | **AZ** | Imprévisible | **Risque Obsolescence**. Make-to-order ou Centralisation. |
        | **CX** | Stable | Stockage de masse (C'est pas cher). |
        | **CZ** | Imprévisible | Stock de sécurité élevé pour éviter les irritants. |
        """)

    st.markdown("#### 📋 Bibliothèque de Formules")
    formulas = [
        {"name": "Wilson (EOQ)", "code": "=SQRT((2*Demande*Cout_Commande)/Cout_Stockage)", "desc": "Quantité optimale de commande."},
        {"name": "Taux de Rotation", "code": "=Ventes_Annuelles / Stock_Moyen", "desc": "Nombre de fois où le stock est renouvelé."},
        {"name": "Délai de Couverture", "code": "= (Stock_Actuel / Conso_Moyenne_Jour)", "desc": "Autonomie du stock en jours."}
    ]
    for f in formulas:
        st.text(f["name"])
        st.code(f["code"], language="excel")
        st.caption(f["desc"])

# --- HELPERS EXCEL ---
def create_ss_excel(avg_c, std_c, lt, std_lt, sl, ss, z):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculateur SS"
    # (Style simplifiés pour la rapidité)
    ws['B2'] = "OUTIL STOCK DE SÉCURITÉ"
    ws['B4'], ws['C4'] = "Paramètre", "Valeur"
    ws.append(["", "Conso Moy", avg_c])
    ws.append(["", "Ecart-type Conso", std_c])
    ws.append(["", "Delai Moy", lt])
    ws.append(["", "Ecart-type Delai", std_lt])
    ws.append(["", "Z Score", z])
    ws.append(["", "STOCK SÉCURITÉ", ss])
    wb.save(output)
    return output.getvalue()