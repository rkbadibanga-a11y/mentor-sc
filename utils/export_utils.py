# utils/export_utils.py
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib import colors

def create_excel_export(title, data_dict, summary_metrics=None):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Mentor SC"

    # --- Header Branding ---
    ws['A1'] = "MENTOR SC - L'EXPERT SUPPLY CHAIN"
    ws['A1'].font = Font(bold=True, size=14, color="00DFD8")
    ws['A2'] = "Contact : r.k.badibanga@gmail.com"
    ws['A3'] = "Developed with ❤️ by Mentor SC"
    ws['A3'].font = Font(italic=True, size=9)

    # Styles
    title_font = Font(name='Arial', size=16, bold=True, color="1E293B")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws['B5'] = title.upper()
    ws['B5'].font = title_font

    # --- Summary Metrics ---
    if summary_metrics:
        row = 7
        ws.cell(row=row, column=2, value="RÉSULTATS CLÉS").font = Font(bold=True, underline="single")
        row += 1
        for label, val in summary_metrics.items():
            ws.cell(row=row, column=2, value=label).font = Font(bold=True)
            ws.cell(row=row, column=3, value=val).font = Font(bold=True, color="007BFF")
            row += 1
        start_data = row + 2
    else:
        start_data = 7

    # --- Data Table ---
    ws.cell(row=start_data, column=2, value="PARAMÈTRE").fill = header_fill
    ws.cell(row=start_data, column=2).font = header_font
    ws.cell(row=start_data, column=3, value="VALEUR").fill = header_fill
    ws.cell(row=start_data, column=3).font = header_font

    current_row = start_data + 1
    for key, value in data_dict.items():
        ws.cell(row=current_row, column=2, value=key).border = border
        ws.cell(row=current_row, column=3, value=value).border = border
        current_row += 1

    # Ajustement colonnes
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25

    wb.save(output)
    return output.getvalue()

def create_pdf_export(title, data_dict, summary_metrics=None):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # --- Background Decoration ---
    c.setFillColor(colors.hexColor("#0D1117")) # Same as app bg
    c.rect(0, height - 4*cm, width, 4*cm, fill=1, stroke=0)

    # --- Header Branding ---
    c.setFillColor(colors.hexColor("#00DFD8"))
    c.setFont("Helvetica-Bold", 22)
    c.drawString(1.5*cm, height - 1.8*cm, "MENTOR SC")
    
    c.setFillColor(colors.white)
    c.setFont("Helvetica", 10)
    c.drawString(1.5*cm, height - 2.5*cm, "L'Expert Supply Chain au service de votre performance")
    c.drawString(1.5*cm, height - 3.0*cm, "Contact : r.k.badibanga@gmail.com")
    
    # White line
    c.setStrokeColor(colors.white)
    c.setLineWidth(1)
    c.line(1.5*cm, height - 3.3*cm, width - 1.5*cm, height - 3.3*cm)

    # --- Report Title ---
    c.setFillColor(colors.hexColor("#1E293B"))
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(width/2, height - 5.5*cm, title)

    # --- Summary Box ---
    y = height - 7.5*cm
    if summary_metrics:
        # Gray background box for summary
        c.setFillColor(colors.hexColor("#F1F5F9"))
        box_height = (len(summary_metrics) * 0.8 + 1) * cm
        c.roundRect(1.5*cm, y - box_height + 0.5*cm, width - 3*cm, box_height, 10, fill=1, stroke=0)
        
        c.setFillColor(colors.hexColor("#1E293B"))
        c.setFont("Helvetica-Bold", 13)
        c.drawString(2*cm, y, "INDICATEURS CALCULÉS :")
        y -= 1*cm
        
        for label, val in summary_metrics.items():
            c.setFont("Helvetica", 11)
            c.drawString(2.5*cm, y, f"{label} :")
            c.setFont("Helvetica-Bold", 12)
            c.setFillColor(colors.hexColor("#007BFF"))
            c.drawRightString(width - 2.5*cm, y, str(val))
            c.setFillColor(colors.hexColor("#1E293B"))
            y -= 0.8*cm
        y -= 1.5*cm

    # --- Table Header ---
    c.setFillColor(colors.hexColor("#1E293B"))
    c.rect(1.5*cm, y, width - 3*cm, 1*cm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y + 0.35*cm, "PARAMÈTRE")
    c.drawString(11*cm, y + 0.35*cm, "VALEUR SAISIE / CALCULÉE")
    
    # --- Table Body ---
    y -= 1*cm
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)
    
    alt = False
    for key, value in data_dict.items():
        if y < 3*cm: # New page if needed
            c.showPage()
            y = height - 3*cm
        
        if alt:
            c.setFillColor(colors.hexColor("#F8FAFC"))
            c.rect(1.5*cm, y - 0.2*cm, width - 3*cm, 0.8*cm, fill=1, stroke=0)
        
        c.setFillColor(colors.black)
        c.drawString(2*cm, y + 0.1*cm, str(key))
        c.drawString(11*cm, y + 0.1*cm, str(value))
        
        # Border
        c.setLineWidth(0.2)
        c.setStrokeColor(colors.grey)
        c.line(1.5*cm, y - 0.2*cm, width - 1.5*cm, y - 0.2*cm)
        
        y -= 0.8*cm
        alt = not alt

    # --- Footer ---
    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.grey)
    c.drawCentredString(width/2, 1.5*cm, "Developed with ❤️ by Mentor SC - r.k.badibanga@gmail.com")
    c.drawCentredString(width/2, 1.1*cm, "Rapport généré automatiquement par l'application Mentor SC")

    c.save()
    buffer.seek(0)
    return buffer.getvalue()